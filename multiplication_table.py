#! python3
""" Multiplication Table Maker
	program takes a number N from the command line and creates an N×N multiplication 
	table in an Excel spreadsheet. The program is run like this:

	python multiplicationTable.py 6
"""

import openpyxl
from openpyxl.styles import Font
import sys
#from openpyxl.cell import get_column_letter, column_index_from_string

def EXCEL_multiplicator(n):
	num = int(n) + 1
	wb = openpyxl.Workbook() # new blank excel object
	sheet = wb.get_sheet_by_name('Sheet')
	font_bold = Font(size=14, bold=True)

	for i in range(1, num): 
		sheet.cell(row=i+1, column=1).value = i
		sheet.cell(row=i+1, column=1).font = font_bold
		sheet.cell(row=1, column=i+1).value = i
		sheet.cell(row=1, column=i+1).font = font_bold

	for row in range(1, num):
		for column in range(1, num):
			sheet.cell(row=row+1, column=column+1).value = row*column	

	wb.save('multiplication_by_{}.xlsx'.format(n))


if __name__ == '__main__':	
	args = sys.argv[1:]

	if not args:
		print('--usage: int')
		sys.exit(1)
	else:
		EXCEL_multiplicator(sys.argv[1])
